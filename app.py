import json
import os
import flask
from werkzeug.utils import secure_filename  # нужна ли вообще?
from openpyxl import load_workbook
import convert_func
import do_xml
import do_xml_from_xlsx


UPLOAD_FOLDER = f'{os.getcwd()}\\uploads'
ALLOWED_EXTENSIONS = set(['xlsx'])

app = flask.Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 1 * 1024 * 1024

def response(code, data):
    return flask.Response(
        status = code,
        mimetype = "application/json",
        response = json.dumps(data) + "\n"
    )

@app.route('/')
def main():
    return flask.render_template('home.html')

@app.route('/', methods=['GET'])
def main_page():
    return response(200, {"status": "OK"})

@app.route('/converter')
def converter():
    id_code = flask.request.args.get('id_code')
    if id_code:
        output_1 = convert_func.convert_to_PC(id_code)
    else:
        output_1 = 'Введите идентификационный код'
    
    product_code = flask.request.args.get('product_code')
    if product_code:
        output_2 = convert_func.convert_to_IC(product_code)
    else:
        output_2 = 'Введите продуктовый код'
    return flask.render_template('converter.html', product_code=output_1, id_code=output_2)

@app.route('/simple_order', methods=['GET'])
def simple_order():
    return flask.render_template('simple_order.html')

@app.route('/order_from_file', methods=['GET'])
def order_from_file():
    return flask.render_template('order_from_file.html')

@app.route('/user_input', methods=['POST'])
def user_input():
    order_dict = flask.request.form.to_dict(flat=False)
    print(order_dict)
    do_xml.xml_builder(order_dict)
    path = "answer.xls"
    return flask.send_file(path, as_attachment=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_xlsx(filepath):
    wb = load_workbook(filepath)
    ws = wb['Данные']
    all_gtin = ws['B']
    all_gtin_list = []
    all_quantity = ws['S']
    all_quantity_list = []
    for row in all_gtin[6:-1]:
        all_gtin_list.append(str(row.value))
    for row in all_quantity[6:-1]:
        all_quantity_list.append(str(row.value))
    return [all_gtin_list, all_quantity_list]

def separate_list(lst, n):
    '''разделяет список на двумерный список в каждом елементе которого n элементов ([1,2,3,4] --> [[1,2],[3,4]] при n=2) '''
    return [lst[i:i + n] for i in range(0, len(lst), n)]

@app.route('/user_file', methods=['POST'])
def user_file():
    order_dict = flask.request.form.to_dict(flat=False)
    print(order_dict)
    file = flask.request.files['file']
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        gtin_and_quantity_lists = parse_xlsx(filepath)
        print(len(gtin_and_quantity_lists[0]))
        print(len(gtin_and_quantity_lists[1]))
        
        # разделяем списки с gtin и quantity на мини списки (до 10 элементов в списке)
        # это нужно, из этих значений далее мы будем формировать "products" в xml 
        # максимально в "products" может быть 10 "product"
        separate_gtin_list = separate_list(gtin_and_quantity_lists[0], 10)
        # print(separate_gtin_list)
        separate_quantity_list = separate_list(gtin_and_quantity_lists[1], 10)
        # print(separate_quantity_list)
        # запаковываем в пары: [список из 10 gtin]:[список из 10 quantity]
        ten_gtin_and_quantity_pack = zip(separate_gtin_list, separate_quantity_list)

        # циклом проходим по zip'у
        c = 0
        for gtin, quantity in ten_gtin_and_quantity_pack:
            copy_order_dict = order_dict.copy()
            copy_order_dict['gtin'] = gtin
            copy_order_dict['quantity'] = quantity
            do_xml_from_xlsx.some_xml_builder(copy_order_dict)
            c += 1
        print(c)
        path = "answer.zip"
        return flask.send_file(path, as_attachment=True)
    return '404'


if __name__ == '__main__':
    app.run(debug=True)